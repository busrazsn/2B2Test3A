from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from constants import globalConstants
import pytest
import openpyxl
import time


class Test_Sauce_Demo:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window() 
        self.driver.get("https://www.saucedemo.com/")

    def teardown_method(self):
        self.driver.quit()

    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx") 
        sheet = excelFile["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data
    
    @pytest.mark.parametrize("username,password", getData())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)
        passwordInput =WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)
        loginButton = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def test_empty_username_and_password_boxes(self):
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.ID,"login-button")))
        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//h3[@data-test='error']")
        assert errorMessage.text == "Epic sadface: Username is required"

    def test_empty_password_box(self,username):
        usernameInput = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID, globalConstants.username_id)))
        usernameInput.send_keys(username)
        loginButton = WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.ID, globalConstants.login_button_id)))
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//h3[@data-test='error']")
        assert errorMessage.text == "Epic sadface: Password is required"
       
    @pytest.mark.dependency(name="valid_login")
    def test_valid_login(self):
        username = self.driver.find_element(By.ID, globalConstants.username_id)
        username.send_keys("standard_user")
        password = self.driver.find_element(By.ID, globalConstants.password_id)
        password.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID, globalConstants.login_button_id)
        loginButton.click()
        time.sleep(5)
        assert "inventory.html" in self.driver.current_url, "Giriş başarılı."
        print("Giriş başarısız.")

    @pytest.mark.dependency (depends=["valid_login"])  
    def test_listed_products(self):
        products = self.driver.find_elements(By.CLASS_NAME,"inventory_item") 
        assert len(products) == 6, "Hata: Toplamda 6 ürün listelenmedi."
        print("Toplamda 6 ürün listelendi.")

    @pytest.mark.dependency (depends=["valid_login"])
    def test_shopping_cart(self):
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[normalize-space()='carry.allTheThings() with the sleek, streamlined Sly Pack that melds uncompromising style with unequaled laptop and tablet protection.']")))
        addToCartButton = self.driver.find_element(By.ID,"add-to-cart-sauce-labs-backpack")
        addToCartButton.click()
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, "//a[@class='shopping_cart_link']")))
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, "//span[@class='shopping_cart_badge']")))
        shoppingCartLink = self.driver.find_element(By.XPATH, "//a[@class='shopping_cart_link']")
        shoppingCartLink.click()
        WebDriverWait(self.driver, 5).until(expected_conditions.url_to_be("https://www.saucedemo.com/cart.html"))
        expected_url = "https://www.saucedemo.com/cart.html"
        assert self.driver.current_url == expected_url, "Sepete gitme işlemi başarısız."
        print("Sepete gitme işlemi başarılı.")
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.XPATH, "//div[@class='cart_item_label']")))
        cartItem = self.driver.find_elements(By.XPATH, "//div[@class='cart_item_label']")
        assert len(cartItem) == 0, "Sepet dolu."
        print("Sepet boş.")
      
        